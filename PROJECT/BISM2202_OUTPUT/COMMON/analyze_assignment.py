#!/usr/bin/env python3
"""Independent validation baseline for BISM2202 Power BI Assessment.

The source workbook is read-only. A typography-only normalization is applied in
memory so the two spellings Marco's Pizza / Marco’s Pizza are treated as one
restaurant in Python and (per the beginner guide) in Power Query.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import platform
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_XLSX_NAME = "Assignment1_BISM2202_pizza_sell_data.xlsx"


def discover_source_xlsx() -> Path:
    """Find the supplied workbook on macOS, Windows, or in the handoff package."""
    configured = os.environ.get("BISM2202_SOURCE_XLSX")
    script = Path(__file__).resolve()
    roots = [Path.cwd(), script.parent, *script.parents, Path.home() / "Downloads"]
    candidates = ([Path(configured).expanduser()] if configured else []) + [
        root / relative
        for root in roots
        for relative in (
            SOURCE_XLSX_NAME,
            Path("INPUTS") / SOURCE_XLSX_NAME,
            Path("input") / SOURCE_XLSX_NAME,
        )
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()
    return Path(SOURCE_XLSX_NAME)


DEFAULT_INPUT = discover_source_xlsx()
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
REQUIRED_FIELDS = [
    "Order ID", "Restaurant Name", "Location", "Order Time", "Delivery Time",
    "Delivery Duration (min)", "Pizza Size", "Pizza Type", "Toppings Count",
    "Distance (km)", "Traffic Level", "Payment Method", "Is Peak Hour",
    "Is Weekend", "Delivery Efficiency (min/km)", "Topping Density",
    "Order Month", "Payment Category", "Estimated Duration (min)",
    "Delay (min)", "Is Delayed", "Pizza Complexity", "Traffic Impact",
    "Order Hour", "Restaurant Avg Time",
]
NUMERIC_FIELDS = [
    "Delivery Duration (min)", "Toppings Count", "Distance (km)",
    "Delivery Efficiency (min/km)", "Topping Density",
    "Estimated Duration (min)", "Delay (min)", "Pizza Complexity",
    "Traffic Impact", "Order Hour", "Restaurant Avg Time",
]
BOOLEAN_FIELDS = ["Is Peak Hour", "Is Weekend", "Is Delayed"]


def json_safe(value: Any) -> Any:
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        value = float(value)
        return None if math.isnan(value) else value
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if pd.isna(value):
        return None
    return value


def records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    return [
        {str(k): json_safe(v) for k, v in row.items()}
        for row in frame.to_dict(orient="records")
    ]


def question_meta(question: str, metric: str, dimension: str, aggregation: str,
                  filter_text: str = "None", sort: str = "None",
                  top_bottom: str = "None", notes: str = "") -> dict[str, str]:
    return {
        "Question": question,
        "Metric": metric,
        "Dimension": dimension,
        "Aggregation": aggregation,
        "Filter": filter_text,
        "Sort": sort,
        "Top N / Bottom N": top_bottom,
        "Notes": notes,
    }


def load_and_validate(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    xls = pd.ExcelFile(path)
    if not xls.sheet_names:
        raise ValueError("The workbook contains no worksheets.")
    raw = pd.read_excel(path, sheet_name=xls.sheet_names[0])

    missing_fields = [c for c in REQUIRED_FIELDS if c not in raw.columns]
    if missing_fields:
        raise ValueError(f"Missing required fields: {missing_fields}")

    df = raw.copy()
    for col in ["Order Time", "Delivery Time"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in NUMERIC_FIELDS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Data-quality correction used by every restaurant-level result.
    restaurant_normalization = {"Marco’s Pizza": "Marco's Pizza"}
    affected = int(df["Restaurant Name"].isin(restaurant_normalization).sum())
    df["Restaurant Name"] = df["Restaurant Name"].replace(restaurant_normalization)

    duplicate_order_rows = int(raw["Order ID"].duplicated(keep=False).sum())
    duplicate_order_values = int(raw.loc[
        raw["Order ID"].duplicated(keep=False), "Order ID"
    ].nunique())
    duplicate_full_rows = int(raw.duplicated().sum())

    duration_from_times = (
        (df["Delivery Time"] - df["Order Time"]).dt.total_seconds() / 60
    )
    checks = {
        "order_month_matches_order_time": int(
            (df["Order Month"] != df["Order Time"].dt.month_name()).sum()
        ),
        "order_hour_matches_order_time": int(
            (df["Order Hour"] != df["Order Time"].dt.hour).sum()
        ),
        "delivery_duration_matches_timestamps": int(
            (~np.isclose(duration_from_times, df["Delivery Duration (min)"], equal_nan=False)).sum()
        ),
        "delay_matches_actual_minus_estimated": int(
            (~np.isclose(
                df["Delivery Duration (min)"] - df["Estimated Duration (min)"],
                df["Delay (min)"], equal_nan=False,
            )).sum()
        ),
        "delivery_efficiency_matches_duration_per_km": int(
            (~np.isclose(
                df["Delivery Duration (min)"] / df["Distance (km)"],
                df["Delivery Efficiency (min/km)"], equal_nan=False,
            )).sum()
        ),
        "topping_density_matches_toppings_per_km": int(
            (~np.isclose(
                df["Toppings Count"] / df["Distance (km)"],
                df["Topping Density"], equal_nan=False,
            )).sum()
        ),
        "restaurant_avg_time_matches_group_average": int(
            (~np.isclose(
                df.groupby("Restaurant Name")["Delivery Duration (min)"].transform("mean"),
                df["Restaurant Avg Time"], equal_nan=False,
            )).sum()
        ),
        "delivery_before_order": int((df["Delivery Time"] < df["Order Time"]).sum()),
        "is_delayed_disagrees_with_dictionary_rule_delay_gt_zero": int(
            (df["Is Delayed"] != (df["Delay (min)"] > 0)).sum()
        ),
        "is_delayed_disagrees_with_actual_rule_duration_gt_30": int(
            (df["Is Delayed"] != (df["Delivery Duration (min)"] > 30)).sum()
        ),
    }

    quality = {
        "source_path": str(path.resolve()),
        "sheet_names": xls.sheet_names,
        "analysis_sheet": xls.sheet_names[0],
        "rows": int(len(raw)),
        "columns": int(len(raw.columns)),
        "field_names": list(raw.columns),
        "missing_required_fields": missing_fields,
        "dtypes_raw": {c: str(t) for c, t in raw.dtypes.items()},
        "null_counts": {c: int(v) for c, v in raw.isna().sum().items()},
        "unique_counts_including_null": {
            c: int(raw[c].nunique(dropna=False)) for c in raw.columns
        },
        "duplicate_order_id_rows": duplicate_order_rows,
        "duplicate_order_id_values": duplicate_order_values,
        "duplicate_full_rows": duplicate_full_rows,
        "boolean_values": {
            c: {str(k): int(v) for k, v in raw[c].value_counts(dropna=False).items()}
            for c in BOOLEAN_FIELDS
        },
        "date_ranges": {
            c: {"min": json_safe(df[c].min()), "max": json_safe(df[c].max()),
                "unparseable": int(df[c].isna().sum())}
            for c in ["Order Time", "Delivery Time"]
        },
        "order_year_counts": {
            str(int(k)): int(v)
            for k, v in df["Order Time"].dt.year.value_counts().sort_index().items()
        },
        "numeric_summary": {
            c: {k: json_safe(v) for k, v in df[c].describe().items()}
            for c in NUMERIC_FIELDS
        },
        "consistency_checks_mismatch_counts": checks,
        "restaurant_name_normalization": {
            "mapping": restaurant_normalization,
            "rows_affected": affected,
            "applied_in_memory_only": True,
        },
        "material_findings": [
            "No missing values, duplicate Order IDs, duplicate rows, unparseable dates, or negative/zero distances were found.",
            "The source contains 188 orders in 2026 even though the scenario says 2024–2025. All 1,004 rows are retained because no question authorizes a year filter.",
            "Three rows use the typographic variant Marco’s Pizza. It is normalized to Marco's Pizza in memory and must be replaced the same way in Power Query.",
            "Is Delayed exactly matches Delivery Duration (min) > 30, not the dictionary rule Delay (min) > 0. None of Q1–Q20 uses Is Delayed.",
            "Topping Density exactly equals Toppings Count / Distance (km), while the dictionary describes a relationship to pizza size. Q18 uses the supplied field as instructed.",
            "Revenue is mentioned in the scenario but no revenue field exists and none of the 20 questions asks for revenue.",
        ],
    }
    return raw, df, quality


def compute_questions(df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    total = int(df["Order ID"].nunique())
    q: dict[str, dict[str, Any]] = {}

    def add(key: str, meta: dict[str, str], frame: pd.DataFrame) -> None:
        q[key] = {"metadata": meta, "result": records(frame)}

    g = (df.groupby("Location")["Order ID"].nunique().rename("Order Count")
         .sort_values(ascending=False).reset_index())
    g["Rank"] = g["Order Count"].rank(method="dense", ascending=False).astype(int)
    add("Q01", question_meta("Top 20 locations based on count of order number",
        "Order Count", "Location", "DISTINCTCOUNT(Order ID)", sort="Order Count descending",
        top_bottom="Top 20", notes="Ties are ordered by Location for deterministic output."),
        g.sort_values(["Order Count", "Location"], ascending=[False, True]).head(20).reset_index(drop=True))

    g = (df.groupby("Pizza Size")["Delivery Duration (min)"].mean()
         .rename("Avg Delivery Duration").reset_index())
    add("Q02", question_meta("Average Delivery Duration by Pizza Size",
        "Avg Delivery Duration", "Pizza Size", "AVERAGE(Delivery Duration (min))"), g)

    g = (df.groupby("Pizza Size")["Order ID"].nunique().rename("Order Count").reset_index())
    g["Share of Orders"] = g["Order Count"] / total
    add("Q03", question_meta("Share of orders by Pizza Size", "Order Count; Share of Orders",
        "Pizza Size", "DISTINCTCOUNT(Order ID); count / total"), g)

    g = (df.groupby("Payment Method")["Order ID"].nunique().rename("Order Count")
         .sort_values(ascending=False).reset_index())
    add("Q04", question_meta("Orders by Payment Method", "Order Count", "Payment Method",
        "DISTINCTCOUNT(Order ID)", sort="Order Count descending"), g)

    g = (df.groupby("Traffic Level")["Order ID"].nunique().rename("Order Count").reset_index())
    g["Share of Orders"] = g["Order Count"] / total
    add("Q05", question_meta("Share of orders by Traffic Level", "Order Count; Share of Orders",
        "Traffic Level", "DISTINCTCOUNT(Order ID); count / total"), g)

    temp = df.assign(**{"Weekend Label": np.where(df["Is Weekend"], "Weekend", "Weekday")})
    g = (temp.groupby("Weekend Label")["Toppings Count"].mean()
         .rename("Avg Toppings Count").reset_index())
    add("Q06", question_meta("Average Toppings Count for weekend vs weekday",
        "Avg Toppings Count", "Weekend Label", "AVERAGE(Toppings Count)"), g)

    g = (df.groupby("Location")["Delivery Duration (min)"].mean()
         .rename("Avg Delivery Duration").sort_values(ascending=False).reset_index())
    g["Rank"] = g["Avg Delivery Duration"].rank(method="dense", ascending=False).astype(int)
    add("Q07", question_meta("Locations with highest average Delivery Duration",
        "Avg Delivery Duration", "Location", "AVERAGE(Delivery Duration (min))",
        sort="Avg Delivery Duration descending", top_bottom="Top 10 for readable chart"),
        g.sort_values(["Avg Delivery Duration", "Location"], ascending=[False, True]).reset_index(drop=True))

    temp = df.assign(**{"Peak Hour Label": np.where(df["Is Peak Hour"], "Peak Hour", "Non-Peak Hour")})
    g = (temp.groupby("Peak Hour Label")["Order ID"].nunique().rename("Order Count").reset_index())
    add("Q08", question_meta("Peak Hour vs Non-Peak order count", "Order Count",
        "Peak Hour Label", "DISTINCTCOUNT(Order ID)"), g)

    temp = df.assign(**{"Month Number": df["Order Time"].dt.month})
    g = (temp.groupby(["Month Number", "Order Month"])["Order ID"].nunique()
         .rename("Order Count").reset_index().sort_values("Month Number"))
    add("Q09", question_meta("Order volume trend across Order Month", "Order Count",
        "Order Month", "DISTINCTCOUNT(Order ID)", sort="Month Number ascending",
        notes="Aggregated across all years per wording; source covers 2024-01 to 2026-07."), g)

    g = (df.groupby("Pizza Size")["Toppings Count"].mean()
         .rename("Avg Toppings Count").reset_index())
    add("Q10", question_meta("Average Toppings Count by Pizza Size", "Avg Toppings Count",
        "Pizza Size", "AVERAGE(Toppings Count)"), g)

    g = (df.groupby("Restaurant Name")["Delay (min)"].mean().rename("Avg Delay")
         .sort_values().reset_index())
    add("Q11", question_meta("Two restaurants with lowest average Delay", "Avg Delay",
        "Restaurant Name", "AVERAGE(Delay (min))", sort="Avg Delay ascending",
        top_bottom="Bottom 2", notes="Restaurant typography normalized consistently."), g.head(2))

    g = (df.groupby("Pizza Type")["Delivery Duration (min)"].mean()
         .rename("Avg Delivery Duration").sort_values(ascending=False).reset_index())
    add("Q12", question_meta("Five Pizza Types with highest average Delivery Duration",
        "Avg Delivery Duration", "Pizza Type", "AVERAGE(Delivery Duration (min))",
        sort="Avg Delivery Duration descending", top_bottom="Top 5"), g.head(5))

    g = (df.groupby("Order Hour")["Order ID"].nunique().rename("Order Count")
         .reset_index().sort_values("Order Hour"))
    add("Q13", question_meta("Order volume by Order Hour", "Order Count", "Order Hour",
        "DISTINCTCOUNT(Order ID)", sort="Order Hour ascending"), g)

    order = pd.CategoricalDtype(["Low", "Medium", "High"], ordered=True)
    temp = df.assign(**{"Traffic Level": df["Traffic Level"].astype(order)})
    g = (temp.groupby("Traffic Level", observed=True)["Delay (min)"].mean()
         .rename("Avg Delay").reset_index())
    g["Traffic Level"] = g["Traffic Level"].astype(str)
    add("Q14", question_meta("Average Delay by Traffic Level", "Avg Delay",
        "Traffic Level", "AVERAGE(Delay (min))", sort="Low, Medium, High"), g)

    g = (df.groupby("Pizza Complexity")["Order ID"].nunique().rename("Order Count")
         .reset_index().sort_values("Pizza Complexity"))
    g["Share of Orders"] = g["Order Count"] / total
    add("Q15", question_meta("Proportion of orders by Pizza Complexity", "Order Count; Share of Orders",
        "Pizza Complexity", "DISTINCTCOUNT(Order ID); count / total", sort="Pizza Complexity ascending",
        notes="The supplied category is numeric and is displayed as a category."), g)

    g = (df.groupby("Restaurant Name").agg(
        **{"Order Count": ("Order ID", "nunique"),
           "Avg Delivery Duration": ("Delivery Duration (min)", "mean"),
           "Avg Delay": ("Delay (min)", "mean")}).reset_index()
         .sort_values("Avg Delivery Duration"))
    g["Delivery Duration Status"] = "Middle"
    g["Delay Status"] = "Middle"
    g.loc[g["Avg Delivery Duration"].idxmin(), "Delivery Duration Status"] = "Fastest"
    g.loc[g["Avg Delivery Duration"].idxmax(), "Delivery Duration Status"] = "Slowest"
    g.loc[g["Avg Delay"].idxmin(), "Delay Status"] = "Lowest"
    g.loc[g["Avg Delay"].idxmax(), "Delay Status"] = "Highest"
    add("Q16", question_meta("Average Delivery Duration and Delay by Restaurant Name",
        "Avg Delivery Duration; Avg Delay", "Restaurant Name", "AVERAGE for both metrics",
        sort="Avg Delivery Duration ascending", notes="Use Fx > Format by field value in Power BI."), g)

    g = (df.groupby("Order Hour").agg(
        **{"Order Count": ("Order ID", "nunique"), "Avg Delay": ("Delay (min)", "mean")})
         .reset_index().sort_values("Order Hour"))
    add("Q17", question_meta("Order volume by Order Hour alongside average Delay",
        "Order Count; Avg Delay", "Order Hour", "DISTINCTCOUNT(Order ID); AVERAGE(Delay (min))",
        sort="Order Hour ascending"), g)

    g = (df.groupby("Pizza Type").agg(
        **{"Order Count": ("Order ID", "nunique"), "Avg Topping Density": ("Topping Density", "mean")})
         .reset_index().sort_values("Order Count", ascending=False))
    add("Q18", question_meta("Average Topping Density by Pizza Type alongside order volume",
        "Avg Topping Density; Order Count", "Pizza Type",
        "AVERAGE(Topping Density); DISTINCTCOUNT(Order ID)", sort="Order Count descending",
        notes="Uses supplied Topping Density field; its actual formula is toppings per km."), g)

    g = (df.groupby(["Payment Method", "Traffic Level"])["Order ID"].nunique()
         .rename("Order Count").reset_index())
    g["Percentage Within Payment Method"] = (
        g["Order Count"] / g.groupby("Payment Method")["Order Count"].transform("sum")
    )
    g = g.sort_values(["Payment Method", "Traffic Level"])
    add("Q19", question_meta("Traffic Level percentage within Payment Method, filterable by Order Time",
        "Order Count; Percentage Within Payment Method", "Payment Method; Traffic Level",
        "DISTINCTCOUNT(Order ID); share within each Payment Method", filter_text="Order Time slicer (Between)",
        notes="Default full-range percentages; each payment method sums to 100%."), g)

    q["Q20_support"] = {
        "metadata": question_meta("Three coordinated visuals with conditional formatting/colors",
            "Reuses Q01/Q12/Q14/Q16/Q17/Q19 metrics", "Restaurant; Traffic; Hour; Location; Pizza Type; Payment",
            "No new aggregation", notes="Design question; source values are the validated earlier results."),
        "result": [],
    }
    return q


def write_quality_report(path: Path, quality: dict[str, Any]) -> None:
    lines = [
        "# BISM2202 Data Quality Report", "",
        f"- Source: `{quality['source_path']}`",
        f"- Sheet(s): {', '.join(quality['sheet_names'])}",
        f"- Analysis sheet: `{quality['analysis_sheet']}`",
        f"- Data rows: {quality['rows']:,}",
        f"- Columns: {quality['columns']}",
        f"- Duplicate Order ID rows: {quality['duplicate_order_id_rows']}",
        f"- Duplicate complete rows: {quality['duplicate_full_rows']}",
        f"- Missing required fields: {quality['missing_required_fields'] or 'None'}", "",
        "## Trust decision", "",
        "The dataset is usable for Q1–Q20 with one documented in-memory text normalization. No source rows are deleted or edited. Restaurant-level calculations normalize the typographic apostrophe in `Marco’s Pizza` to `Marco's Pizza`; the Power BI guide applies the same Power Query replacement so Python and Power BI use one grain.", "",
        "## Material findings", "",
    ]
    lines.extend([f"- {x}" for x in quality["material_findings"]])
    lines += ["", "## Fields, types, nulls, and unique values", "",
              "| Field | Raw dtype | Nulls | Unique (incl. null) |", "|---|---:|---:|---:|"]
    for field in quality["field_names"]:
        lines.append(
            f"| {field} | {quality['dtypes_raw'][field]} | {quality['null_counts'][field]} | "
            f"{quality['unique_counts_including_null'][field]} |"
        )
    lines += ["", "## Boolean fields", ""]
    for field, counts in quality["boolean_values"].items():
        lines.append(f"- `{field}`: {counts}")
    lines += ["", "## Date checks", ""]
    for field, info in quality["date_ranges"].items():
        lines.append(f"- `{field}`: {info['min']} to {info['max']}; unparseable = {info['unparseable']}")
    lines.append(f"- Orders by year: {quality['order_year_counts']}")
    lines += ["", "## Consistency checks (mismatch counts)", ""]
    for check, count in quality["consistency_checks_mismatch_counts"].items():
        lines.append(f"- `{check}`: {count}")
    lines += ["", "## Numeric ranges", "",
              "| Field | Count | Min | Mean | Max |", "|---|---:|---:|---:|---:|"]
    for field, stats in quality["numeric_summary"].items():
        lines.append(
            f"| {field} | {stats['count']:.0f} | {stats['min']:.6g} | "
            f"{stats['mean']:.6g} | {stats['max']:.6g} |"
        )
    lines += ["", "## Analytical implications", "",
              "- Use `DISTINCTCOUNT(Order ID)` for all order counts; `Order ID` is unique across all 1,004 rows.",
              "- Do not filter out 2026 unless the instructor explicitly requests a 2024–2025-only correction.",
              "- Use the actual `Order Time` month number to sort `Order Month` January through December.",
              "- Treat `Pizza Complexity` as a categorical grouping even though its source type is numeric.",
              "- Do not use `Is Delayed` to answer delay questions; Q11/Q14/Q16/Q17 use the numeric `Delay (min)` field.",
              "- Q18 uses the supplied `Topping Density` field and reports the definition mismatch as a caveat, without inventing a replacement formula."]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_excel(path: Path, quality: dict[str, Any], questions: dict[str, dict[str, Any]]) -> None:
    summary_rows = []
    for key, payload in questions.items():
        meta = payload["metadata"]
        summary_rows.append({"Question ID": key, **meta, "Result Rows": len(payload["result"])})
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        for key, payload in questions.items():
            result = pd.DataFrame(payload["result"])
            meta = pd.DataFrame([payload["metadata"]])
            meta.to_excel(writer, sheet_name=key, index=False, startrow=0)
            result.to_excel(writer, sheet_name=key, index=False, startrow=3)
        quality_rows = []
        for field in quality["field_names"]:
            quality_rows.append({
                "Field": field, "Raw dtype": quality["dtypes_raw"][field],
                "Nulls": quality["null_counts"][field],
                "Unique (incl. null)": quality["unique_counts_including_null"][field],
            })
        pd.DataFrame(quality_rows).to_excel(writer, sheet_name="Data_Quality", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        sub_fill = PatternFill("solid", fgColor="D9EAF7")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2" if ws.title in ("Summary", "Data_Quality") else "A5"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            if ws.title.startswith("Q"):
                for cell in ws[4]:
                    cell.fill = sub_fill
                    cell.font = Font(bold=True, color="1F1F1F")
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, float):
                        if cell.column_letter and "Share" in str(ws.cell(4 if ws.title.startswith('Q') else 1, cell.column).value):
                            cell.number_format = "0.00%"
                        elif "Percentage" in str(ws.cell(4 if ws.title.startswith('Q') else 1, cell.column).value):
                            cell.number_format = "0.00%"
                        else:
                            cell.number_format = "0.000000"
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col_idx in range(1, ws.max_column + 1):
                values = [str(ws.cell(r, col_idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
                width = min(max(12, max(len(v) for v in values) + 2), 42)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.sheet_view.showGridLines = False


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    raw, clean, quality = load_and_validate(args.input)
    questions = compute_questions(clean)
    payload = {
        "metadata": {
            "course": "BISM2202",
            "assessment": "Data Visualization Using Microsoft Power BI - Assessment Task",
            "source": str(args.input.resolve()),
            "analysis_grain": "One row per unique Order ID",
            "total_orders": int(clean["Order ID"].nunique()),
            "source_rows": int(len(raw)),
            "python": platform.python_version(),
            "normalization": "Marco’s Pizza -> Marco's Pizza (3 rows, in memory only)",
        },
        "data_quality": quality,
        "questions": questions,
    }
    json_path = args.output_dir / "analysis_results.json"
    xlsx_path = args.output_dir / "analysis_results.xlsx"
    report_path = args.output_dir / "DATA_QUALITY_REPORT.md"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(xlsx_path, quality, questions)
    write_quality_report(report_path, quality)
    print(f"Wrote {json_path}")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {report_path}")
    print(f"Validated {len(raw):,} rows and {clean['Order ID'].nunique():,} distinct orders.")


if __name__ == "__main__":
    main()
